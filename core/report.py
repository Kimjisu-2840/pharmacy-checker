# core/report.py

import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_from_view_df(df, today_str, provider):
    """
    검수 현황 데이터프레임을 바탕으로 하이라이트 색상이 반영된 엑셀 보고서(bytes)를 생성합니다.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "입고검수결과"

    # 1. 메인 타이틀 헤더
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"[{today_str}] {provider} 의약품 입고 검수 보고서"
    title_cell.font = Font(name="맑은 고딕", size=15, bold=True, color="1B365D")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 2. 표 헤더 설정
    headers = ["No", "제품코드", "제품전산출력명", "제품규격", "입고예정수량", "스캔수량", "수량차이", "검수상태"]
    ws.append([]) # 2행 공백
    ws.append(headers) # 3행 헤더
    
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D0D3D4'),
        right=Side(style='thin', color='D0D3D4'),
        top=Side(style='thin', color='D0D3D4'),
        bottom=Side(style='thin', color='D0D3D4')
    )

    ws.row_dimensions[3].height = 26
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 3. 상태별 하이라이트 색상 정의 (Streamlit 대시보드와 1:1 대응)
    fill_unregistered = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid") # 미등록 (소프트 핑크)
    fill_completed    = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid") # 완료 (연두)
    fill_over         = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid") # 초과 (연빨강)
    fill_in_progress  = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid") # 차이/진행중 (노랑)
    fill_unscanned    = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # 미스캔 (흰색)

    # 4. 데이터 행 작성 및 하이라이트 스타일링
    current_row = 4
    for _, row in df.iterrows():
        no = row.get("No", "")
        code = row.get("제품코드", "")
        name = row.get("제품전산출력명", "")
        spec = row.get("제품규격", "")
        expected = int(row.get("수량", 0))
        scanned = int(row.get("스캔수량", 0))
        diff = scanned - expected
        is_registered = row.get("등록여부", True)

        # 상태 판단 및 하이라이트 적용
        if not is_registered:
            status = "미등록"
            row_fill = fill_unregistered
        elif scanned == expected and expected > 0:
            status = "검수완료"
            row_fill = fill_completed
        elif scanned > expected:
            status = f"초과(+{diff})"
            row_fill = fill_over
        elif scanned > 0:
            status = f"차이({diff})"
            row_fill = fill_in_progress
        else:
            status = "미스캔"
            row_fill = fill_unscanned

        row_data = [no, code, name, spec, expected, scanned, diff, status]
        ws.append(row_data)
        ws.row_dimensions[current_row].height = 22

        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.fill = row_fill
            cell.font = Font(name="맑은 고딕", size=10)
            cell.border = thin_border
            
            # 정렬 및 수치 포맷
            if col_num in [1, 2, 7, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num in [5, 6]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        current_row += 1

    # 5. 총 합계 행 추가 (수식 적용)
    ws.append([])
    summary_row = current_row + 1
    ws.cell(row=summary_row, column=3, value="총 합계").font = Font(name="맑은 고딕", size=11, bold=True)
    ws.cell(row=summary_row, column=3).alignment = Alignment(horizontal="center", vertical="center")
    
    sum_expected = ws.cell(row=summary_row, column=5, value=f"=SUM(E4:E{current_row-1})")
    sum_expected.font = Font(name="맑은 고딕", size=11, bold=True)
    sum_expected.number_format = '#,##0'
    sum_expected.alignment = Alignment(horizontal="right", vertical="center")
    
    sum_scanned = ws.cell(row=summary_row, column=6, value=f"=SUM(F4:F{current_row-1})")
    sum_scanned.font = Font(name="맑은 고딕", size=11, bold=True)
    sum_scanned.number_format = '#,##0'
    sum_scanned.alignment = Alignment(horizontal="right", vertical="center")

    # 6. 컬럼 너비 자동 맞춤
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, summary_row]:
                continue
            val_str = str(cell.value or '')
            length = sum(2 if ord(c) > 128 else 1 for c in val_str)
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 메모리 버퍼로 반환
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
